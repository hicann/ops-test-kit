#!/usr/bin/env python3
# -*- coding: UTF-8 -*-
# Copyright (c) 2026 Huawei Technologies Co., Ltd.
# This program is free software, you can redistribute it and/or modify it under the terms and conditions of
# CANN Open Software License Agreement Version 2.0 (the "License").
# Please refer to the License for details. You may not use this file except in compliance with the License.
# THIS SOFTWARE IS PROVIDED ON AN "AS IS" BASIS, WITHOUT WARRANTIES OF ANY KIND, EITHER EXPRESS OR IMPLIED,
# INCLUDING BUT NOT LIMITED TO NON-INFRINGEMENT, MERCHANTABILITY, OR FITNESS FOR A PARTICULAR PURPOSE.
# See LICENSE in the root of the software repository for the full text of the License.
"""
Unified table reader.

Reads CSV and XLSX inputs into a (header, rows) pair of stripped strings,
so downstream testcase parsing is format-agnostic. CSV path mirrors the
original UniversalTestcaseFactory._read_csv semantics; XLSX path coerces
every cell to str (None -> "") to match CSV's text-only behavior.
"""

__all__ = ["read_table", "read_csv_rows", "read_xlsx_rows", "resolved_sheet"]

import csv


def read_csv_rows(fileobj) -> tuple:
    """Read a CSV file object into (header, rows).

    Each cell is stripped; empty lines are dropped; the first non-empty
    line is the header. Matches the legacy _read_csv behavior exactly.
    """
    rows = []
    for row in csv.reader(fileobj):
        row = [column.strip() for column in row]
        if row:
            rows.append(row)
    if not rows:
        raise ValueError("Empty table: no rows found")
    return rows[0], rows[1:]


def read_xlsx_rows(path, sheet=None) -> tuple:
    """Read an .xlsx workbook into (header, rows) of stripped strings.

    sheet: worksheet name; default first worksheet. Cells are coerced to
    str (None -> "") and stripped. Fully-empty rows are dropped, matching
    the CSV reader's empty-line filtering. data_only=True reads cached
    formula values; numeric cells become their decimal string form, so
    shape/dtype columns should be formatted as Text in the workbook.
    """
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover
        raise ImportError("openpyxl is required to read .xlsx inputs; install it with `pip install openpyxl`") from exc

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet is not None:
            if sheet not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet}' not found in {path}; available: {wb.sheetnames}")
            ws = wb[sheet]
        else:
            ws = wb[wb.sheetnames[0]]
        rows = []
        for raw in ws.iter_rows(values_only=True):
            row = [("" if cell is None else str(cell)).strip() for cell in raw]
            if any(cell != "" for cell in row):
                rows.append(row)
        if not rows:
            sheet_name = sheet or wb.sheetnames[0]
            raise ValueError(f"Empty sheet '{sheet_name}' in {path}")
        return rows[0], rows[1:]
    finally:
        wb.close()


def read_table(path, sheet=None) -> tuple:
    """Dispatch by file suffix: .xlsx/.xlsm -> openpyxl, otherwise CSV."""
    if path.lower().endswith((".xlsx", ".xlsm")):
        return read_xlsx_rows(path, sheet)
    with open(path, newline="", encoding="utf-8") as f:
        return read_csv_rows(f)


def resolved_sheet(path, sheet=None):
    """Return the worksheet name that will actually be read.

    A specified sheet is returned as-is. For xlsx without an explicit
    sheet, returns the first worksheet name (the default read target).
    For csv, returns None (no worksheet concept).
    """
    if not path.lower().endswith((".xlsx", ".xlsm")):
        return None
    if sheet is not None:
        return sheet
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover
        raise ImportError("openpyxl is required to read .xlsx inputs; install it with `pip install openpyxl`") from exc
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        return wb.sheetnames[0] if wb.sheetnames else None
    finally:
        wb.close()
